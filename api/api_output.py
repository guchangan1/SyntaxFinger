# -*- coding: utf-8 -*-

"""
@Time : 2023/9/9
@Author : guchangan1
"""
import os
import time

import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.utils import get_column_letter

from config.data import logging, path, Webinfo, Urlerror


class api_output:
    def __init__(self, projectname):
        self.nowTime = time.strftime("%Y%m%d%H%M%S",time.localtime())
        #excel名称
        self.projectname = projectname + "_" +self.nowTime + ".xlsx"
        # excel的保存路径
        self.path_excel = os.path.join(path.apioutputdir, self.projectname)
        self.excelSavePath = self.path_excel
        self.excel = openpyxl.Workbook()
        self.Sheet_line = 1   # 空间引擎聚合结果表格的行
        self.Sheet2_line = 1  # 指纹识别结果表格的行
        self.Sheet3_line = 1  # 指纹识别失败结果表格的行

    def set_column_widths(sheet, column_width):
        for idx in range(1, column_width + 1):
            column_letter = get_column_letter(idx)
            sheet.column_dimensions[column_letter].width = 15  # 设定宽度为15
    def api_outXls(self, webSpaceResult):
        self.sheet = self.excel.create_sheet(title="空间引擎聚合结果", index=0)  # 创建空间引擎聚合结果工作区
        if self.Sheet_line == 1:
            self.sheet.cell(self.Sheet_line, 1).value = '空间引擎名'
            self.sheet.cell(self.Sheet_line, 2).value = 'url'
            self.sheet.cell(self.Sheet_line, 3).value = 'title'
            self.sheet.cell(self.Sheet_line, 4).value = 'ip'
            self.sheet.cell(self.Sheet_line, 5).value = 'port'
            self.sheet.cell(self.Sheet_line, 6).value = 'server'
            self.sheet.cell(self.Sheet_line, 7).value = 'address'
            self.sheet.cell(self.Sheet_line, 8).value = 'icp_number'
            self.sheet.cell(self.Sheet_line, 9).value = 'icp_name'
            self.sheet.cell(self.Sheet_line, 10).value = 'isp'
            self.sheet.cell(self.Sheet_line, 11).value = '查询语句'
            self.Sheet_line += 1

        for result in webSpaceResult:
            try:
                title = ILLEGAL_CHARACTERS_RE.sub(r'', result["title"])
            except Exception as e:
                title = ''
            self.sheet.cell(self.Sheet_line, 1).value = result["api"]
            self.sheet.cell(self.Sheet_line, 2).value = result["url"]
            self.sheet.cell(self.Sheet_line, 3).value = title
            self.sheet.cell(self.Sheet_line, 4).value = result["ip"]
            self.sheet.cell(self.Sheet_line, 5).value = result["port"]
            self.sheet.cell(self.Sheet_line, 6).value = result["server"]
            self.sheet.cell(self.Sheet_line, 7).value = result["region"]
            self.sheet.cell(self.Sheet_line, 8).value = result["icp"]
            self.sheet.cell(self.Sheet_line, 9).value = result["icp_name"]
            self.sheet.cell(self.Sheet_line, 10).value = result["isp"]
            self.sheet.cell(self.Sheet_line, 11).value = result["syntax"]
            self.Sheet_line += 1

        self.excel.save(self.excelSavePath)
        successMsg = "空间引擎聚合结果成功保存！输出路径为:{0}".format(self.excelSavePath)
        logging.success(successMsg)

    def Succfinger_outXls(self):
        self.sheet2 = self.excel.create_sheet(title="指纹识别结果", index=1)  # 创建指纹识别结果工作区
        if self.Sheet2_line == 1:
            self.sheet2.cell(self.Sheet2_line, 1).value = '评分'
            self.sheet2.cell(self.Sheet2_line, 2).value = 'url'
            self.sheet2.cell(self.Sheet2_line, 3).value = '标题'
            self.sheet2.cell(self.Sheet2_line, 4).value = '状态码'
            self.sheet2.cell(self.Sheet2_line, 5).value = '指纹&CMS'
            self.sheet2.cell(self.Sheet2_line, 6).value = 'Server头'
            self.sheet2.cell(self.Sheet2_line, 7).value = 'ip'
            self.sheet2.cell(self.Sheet2_line, 8).value = '归属地'
            self.sheet2.cell(self.Sheet2_line, 9).value = '是否CDN'
            self.sheet2.cell(self.Sheet2_line, 10).value = 'isp'
            self.sheet2.cell(self.Sheet2_line, 11).value = 'icp'
            self.Sheet2_line += 1
        for vaule in Webinfo.result:
            try:
                title = ILLEGAL_CHARACTERS_RE.sub(r'', vaule["title"])
            except Exception as e:
                title = ''
            self.sheet2.cell(self.Sheet2_line, 1).value = ""
            self.sheet2.cell(self.Sheet2_line, 2).value = vaule["url"]
            self.sheet2.cell(self.Sheet2_line, 3).value = title
            self.sheet2.cell(self.Sheet2_line, 4).value = vaule["status"]
            self.sheet2.cell(self.Sheet2_line, 5).value = vaule["cms"]
            self.sheet2.cell(self.Sheet2_line, 6).value = vaule["Server"]
            self.sheet2.cell(self.Sheet2_line, 7).value = vaule["ip"]
            self.sheet2.cell(self.Sheet2_line, 8).value = vaule["address"]
            self.sheet2.cell(self.Sheet2_line, 9).value = vaule["iscdn"]
            self.sheet2.cell(self.Sheet2_line, 10).value = vaule["isp"]
            self.sheet2.cell(self.Sheet2_line, 11).value = vaule["icp"]
            self.Sheet2_line += 1

        self.excel.save(self.excelSavePath)
        successMsg = "存活指纹识别结果成功保存！输出路径为:{0}".format(self.excelSavePath)
        logging.success(successMsg)

    def Failfinger_outXls(self):
        self.sheet3 = self.excel.create_sheet(title="请求失败结果", index=1)  # 创建指纹识别结果工作区
        if self.Sheet3_line == 1:
            self.sheet3.cell(self.Sheet3_line, 1).value = '评分'
            self.sheet3.cell(self.Sheet3_line, 2).value = 'url'
            self.sheet3.cell(self.Sheet3_line, 3).value = '标题'
            self.sheet3.cell(self.Sheet3_line, 4).value = '状态码'
            self.sheet3.cell(self.Sheet3_line, 5).value = '指纹&CMS'
            self.sheet3.cell(self.Sheet3_line, 6).value = 'Server头'
            self.sheet3.cell(self.Sheet3_line, 7).value = 'ip'
            self.sheet3.cell(self.Sheet3_line, 8).value = '归属地'
            self.sheet3.cell(self.Sheet3_line, 9).value = '是否CDN'
            self.sheet3.cell(self.Sheet3_line, 10).value = 'isp'
            self.sheet3.cell(self.Sheet3_line, 11).value = 'icp'
            self.Sheet3_line += 1
        for vaule in Urlerror.result:
            self.sheet3.cell(self.Sheet3_line, 1).value = ""
            self.sheet3.cell(self.Sheet3_line, 2).value = vaule["url"]
            self.sheet3.cell(self.Sheet3_line, 3).value = vaule["title"]
            self.sheet3.cell(self.Sheet3_line, 4).value = vaule["status"]
            self.sheet3.cell(self.Sheet3_line, 5).value = vaule["cms"]
            self.sheet3.cell(self.Sheet3_line, 6).value = vaule["Server"]
            self.sheet3.cell(self.Sheet3_line, 7).value = vaule["ip"]
            self.sheet3.cell(self.Sheet3_line, 8).value = vaule["address"]
            self.sheet3.cell(self.Sheet3_line, 9).value = vaule["iscdn"]
            self.sheet3.cell(self.Sheet3_line, 10).value = vaule["isp"]
            self.sheet3.cell(self.Sheet3_line, 11).value = vaule["icp"]
            self.Sheet3_line += 1

        self.excel.save(self.excelSavePath)
        successMsg = "请求失败的结果成功保存！输出路径为:{0}".format(self.excelSavePath)
        logging.success(successMsg)